import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

# Conexión a la base de datos en Microsoft Azure
#azure_engine = create_engine('postgresql://)

# Conexión a la base de datos en MITIC
mitic_engine = create_engine(f'postgresql://usuario:contraseña@168.90.177.232:2024/bbdd-monitoreo-cid')

# Obtener los datos relevantes de la base de datos en Azure

# Obtener los datos relevantes de la base de datos MITIC filtrando por troncales
mitic_query = """
SELECT eot, ruta_dec, ruta_hex, sentido, linea, ramal, origen, destino, 
       identificacion, "identificador troncal temporal hasta nueva rev."
FROM catalogo_rutas
WHERE "identificador troncal temporal hasta nueva rev." IN ('Troncal 1', 'Troncal 2')
"""
mitic_df = pd.read_sql(mitic_query, mitic_engine)

# Hacer los cálculos correspondientes con el dataframe


# Guardar los resultados en la base de datos
#No creo que haga falta
#result.to_sql('resultado_frecuencia', mitic_engine, if_exists='replace', index=False)

#Guardar los resultados en un archivo Excel a partir de la fila 10
output_file = 'C:/Users/support/Documents/VMT - CID/Monitoreo/informe_fiduciario_quincenal.xlsx'
try:
    # Cargar el workbook existente
    workbook = load_workbook(output_file)
    sheet = workbook.active
except FileNotFoundError:
    # Si el archivo no existe, crear un nuevo workbook y hoja
    workbook = Workbook()
    sheet = workbook.active

# Convertir el DataFrame a filas
rows = dataframe_to_rows(result, index=False, header=True)

# Escribir los datos a partir de la fila 10
for i, row in enumerate(rows, start=10):
    for j, value in enumerate(row, start=1):
        sheet.cell(row=i, column=j, value=value)

# Guardar el archivo
workbook.save(output_file)

# Enviar el archivo Excel por correo electrónico
def send_email(subject, body, to, attachment):
    msg = MIMEMultipart()
    msg['From'] = 'billetajevmt@gmail.com'
    msg['To'] = to
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'plain'))

    with open(attachment, 'rb') as f:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename={attachment}')
        msg.attach(part)

    with smtplib.SMTP('smtp.gmail.com', 587) as server:
        server.starttls()
        server.login('billetajevmt@gmail.com', 'kvsl hqdo fkvz hnwh')
        server.send_message(msg)

# Definir los parámetros del correo
subject = 'Resultado de Frecuencia'
body = """
[CORREO GENERADO AUTOMÁTICAMENTE]

Buenos días.

Adjunto se encuentra el archivo con el resultado de las métricas solicitadas.

Atentamente
Equipo de la Coordinación de Innovación y Desarrollo
DMT-VMT-MOPC
"""
to = 'lprafael1710@gmail.com'

# Enviar el correo con el archivo adjunto
send_email(subject, body, to, output_file)

print(result)
