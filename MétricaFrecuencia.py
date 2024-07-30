import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

# Conexión a la base de datos en Microsoft Azure
azure_engine = create_engine('postgresql://vmtuser:$$Tr4nsp0rt3$$@flex-opentransit-prod-1.postgres.database.azure.com:5432/opentransit-prod')

# Conexión a la base de datos en MITIC
mitic_engine = create_engine(f'postgresql://cid_admin_user:vmtdmtcidccm@168.90.177.232:2024/bbdd-monitoreo-cid')

# Obtener los datos relevantes de la base de datos en Azure
azure_query = """
SELECT DISTINCT eventos_dia.idsam, eventos_dia.idrutaestacion, eventos_dia.fechahoraevento::date as fecha, extract(hour FROM fechahoraevento) AS hora
FROM (SELECT DISTINCT (c_transacciones.idsam::text || c_transacciones.consecutivoevento::text || c_transacciones.serialmediopago::text) AS evento, c_transacciones.idsam, 
      c_transacciones.serialmediopago, c_transacciones.fechahoraevento, c_transacciones.idproducto, c_transacciones.identidad, c_transacciones.iddispositivo, 
      c_transacciones.montoevento, c_transacciones.tipoevento, c_transacciones.idrutaestacion
     FROM c_transacciones
     WHERE (c_transacciones.fechahoraevento BETWEEN current_date ::timestamp - interval '1 days' and current_date)
       AND (c_transacciones.fechahoraevento::time between '05:00:00' AND '07:59:59' or
            c_transacciones.fechahoraevento::time between '16:00:00' AND '18:59:59')
       AND c_transacciones.TIPOEVENTO in (4,8)
    ) eventos_dia
"""
azure_df = pd.read_sql(azure_query, azure_engine)

# Obtener los datos relevantes de la base de datos MITIC filtrando por troncales
mitic_query = """
SELECT g.gre_nombre as "GREMIO", e.eot_nombre AS "EOT", cr.ruta_dec, cr.ruta_hex, cr."identificador_troncal" AS "Identificador Troncal"
FROM catalogo_rutas cr
    join eots e on cr.id_eot_catalogo=e.cod_catalogo
	join gremios g on e.gre_id=g.gre_id
WHERE "identificador_troncal" IN ('Troncal 1', 'Troncal 2')
"""
mitic_df = pd.read_sql(mitic_query, mitic_engine)

# Convertir ruta_hex e idrutaestacion a cadena de texto y luego a mayúsculas
azure_df['idrutaestacion'] = azure_df['idrutaestacion'].astype(str).str.upper()
mitic_df['ruta_hex'] = mitic_df['ruta_hex'].astype(str).str.upper()

# Unir (machear) los datos basados en las columnas comunes
merged_df = pd.merge(azure_df, mitic_df, how='right', left_on='idrutaestacion', right_on='ruta_hex')

# Calcular la cantidad de buses por EOT y Troncal en las franjas horarias
result = merged_df.groupby(['GREMIO', 'EOT', 'Identificador Troncal']).agg(
    cant_buses_mañana=('idsam', lambda x: x[merged_df['hora'].between(5, 7)].nunique()),
    cant_buses_tarde=('idsam', lambda x: x[merged_df['hora'].between(16, 18)].nunique())
).reset_index()

# Agregar la columna de cumplimiento
result['cumplimiento'] = result.apply(
    lambda row: 'Cumple' if row['cant_buses_mañana'] >= 12 and row['cant_buses_tarde'] >= 12 else 'Incumple',
    axis=1
)

# Ordenar los datos por EOT
result = result.sort_values(by='EOT')

# Guardar los resultados en la base de datos
result.to_sql('resultado_frecuencia', mitic_engine, if_exists='replace', index=False)

#Guardar los resultados en un archivo Excel a partir de la fila 10
output_file = 'C:/Users/support/Documents/VMT - CID/Monitoreo/resultado_frecuencia.xlsx'
try:
    # Cargar el workbook existente
    workbook = load_workbook(output_file)
    sheet = workbook.active
except FileNotFoundError:
    # Si el archivo no existe, crear un nuevo workbook y hoja
    workbook = Workbook()
    sheet = workbook.active

# Convertir el DataFrame a filas
rows = dataframe_to_rows(result, index=False, header=True)

# Escribir los datos a partir de la fila 10
for i, row in enumerate(rows, start=10):
    for j, value in enumerate(row, start=1):
        sheet.cell(row=i, column=j, value=value)

# Guardar el archivo
workbook.save(output_file)

# Enviar el archivo Excel por correo electrónico
def send_email(subject, body, to, attachment):
    msg = MIMEMultipart()
    msg['From'] = 'billetajevmt@gmail.com'
    msg['To'] = to
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'plain'))

    with open(attachment, 'rb') as f:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename={attachment}')
        msg.attach(part)

    with smtplib.SMTP('smtp.gmail.com', 587) as server:
        server.starttls()
        server.login('billetajevmt@gmail.com', 'kvsl hqdo fkvz hnwh')
        server.send_message(msg)

# Definir los parámetros del correo
subject = 'Resultado de Frecuencia del día anterior'
body = """
[CORREO GENERADO AUTOMÁTICAMENTE]

Buenos días.

Adjunto se encuentra el archivo con el resultado de las métricas solicitadas.

Atentamente
Equipo de la Coordinación de Innovación y Desarrollo
DMT-VMT-MOPC
"""
#to = 'lprafael1710@gmail.com'
to = 'lprafael1710@gmail.com, transporte.mopc@gmail.com'
# Enviar el correo con el archivo adjunto
send_email(subject, body, to, output_file)

print(result)
